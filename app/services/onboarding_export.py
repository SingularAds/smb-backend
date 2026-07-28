"""Excel (.xlsx) export of onboarding-funnel prospects for the dashboard.

Builds on ``get_platform_overview`` — the exact data the dashboard funnel
renders (same demo-session exclusion, test filtering, global-number scoping
and deleted-session reconstruction) — so the exported sheet always matches
what the screen shows. The selected export window maps to the funnel's own
window (``funnel_start``/``funnel_end``).

One row per onboarding journey: every registration session active in the
window, plus businesses whose session no longer exists (marked
"Reconstructed"). Deepest funnel stage, business details, pairing state and
acquisition context are included so the sheet is directly usable for
follow-up calls without opening the dashboard.
"""
from __future__ import annotations

import re
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services import analytics_service

# Column spec: (header, row-dict key, minimum width). Widths grow to fit the
# longest cell in the column, capped at _MAX_COL_WIDTH so one long address
# can't blow up the layout.
_COLUMNS: list[tuple[str, str, int]] = [
    ("Owner name",          "ownerName",         16),
    ("Owner phone",         "ownerPhone",        15),
    ("Business name",       "businessName",      20),
    ("Business ID",         "businessId",        22),
    ("Business address",    "address",           28),
    ("Business type",       "businessType",      14),
    ("Onboarding stage",    "stage",             26),
    ("Current step",        "currentStep",       16),
    ("WhatsApp paired",     "whatsappPaired",    15),
    ("Started at (UTC)",    "startedAt",         17),
    ("Business created at", "businessCreatedAt", 19),
    ("Onboarding number",   "onboardingNumber",  17),
    ("Channel",             "channel",           12),
]

_MAX_COL_WIDTH = 50

_HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_CELL_BORDER = Border(bottom=Side(style="thin", color="E3E6EA"))


def _fmt_dt(iso: str | None) -> str:
    """ISO timestamp → 'YYYY-MM-DD HH:MM' (UTC), or '' when absent."""
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(str(iso).replace("Z", "+00:00")).strftime(
            "%Y-%m-%d %H:%M"
        )
    except ValueError:
        return str(iso)


def collect_rows(
    start: datetime,
    end: datetime,
    global_device: str | None = None,
) -> list[dict]:
    """One export row per onboarding journey in [start, end].

    The overview's per-stage funnel lists are cumulative (started ⊇ details ⊇
    onboarded ⊇ paired) and reuse the same entry per prospect, so scanning
    from the deepest stage down, the first list an entry appears in names its
    deepest stage.
    """
    ov = analytics_service.get_platform_overview(
        start, end, global_device=global_device,
        funnel_start=start, funnel_end=end,
    )
    accounts = {a["id"]: a for a in ov["accounts"]}

    rows: list[dict] = []
    seen: set[tuple] = set()
    for st in reversed(ov["funnel"]):
        for entry in st["sessions"]:
            key = (entry.get("phone"), entry.get("businessId"))
            if key in seen:
                continue
            seen.add(key)
            acc = accounts.get(entry.get("businessId") or "") or {}
            has_business = bool(entry.get("businessId"))
            rows.append({
                "ownerName": entry.get("name") or acc.get("ownerName") or "",
                "ownerPhone": entry.get("phone") or "",
                "businessName": entry.get("businessName") or "",
                "businessId": entry.get("businessId") or "",
                "address": acc.get("address") or "",
                "businessType": acc.get("businessType") or "",
                "stage": st["label"],
                "currentStep": entry.get("currentStep") or "",
                # Blank (not "No") when no business exists yet — pairing is
                # simply not applicable to a mid-flow prospect.
                "whatsappPaired": (
                    ("Yes" if acc.get("whatsappPaired") else "No")
                    if has_business else ""
                ),
                "startedAt": _fmt_dt(entry.get("startedAt")),
                "businessCreatedAt": _fmt_dt(acc.get("createdAt")),
                "onboardingNumber": entry.get("onboardingNumber") or "",
                "channel": entry.get("channel") or "",
            })

    rows.sort(key=lambda r: r["startedAt"], reverse=True)
    return rows


def build_workbook(rows: list[dict]) -> bytes:
    """Rows → styled .xlsx bytes: bold header, frozen top row, autofilter,
    per-column widths sized to content."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Onboarding"

    for col_idx, (header, _key, _min_w) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_header, key, _min_w) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key) or "")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = _CELL_BORDER

    # Fit each column to its longest value (header included), capped so long
    # addresses can't distort the sheet.
    for col_idx, (header, key, min_w) in enumerate(_COLUMNS, start=1):
        longest = max(
            [len(header)] + [len(str(r.get(key) or "")) for r in rows]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(min_w, longest + 2), _MAX_COL_WIDTH
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_COLUMNS))}{max(len(rows) + 1, 2)}"
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_onboarding_export(
    start: datetime,
    end: datetime,
    global_device: str | None = None,
) -> tuple[bytes, str]:
    """The complete export: (xlsx bytes, download filename)."""
    rows = collect_rows(start, end, global_device=global_device)
    scope = ""
    if global_device:
        scope = "_" + re.sub(r"[^A-Za-z0-9_-]", "", global_device)
    # `end` is exclusive (a bare to=YYYY-MM-DD resolves to next-day midnight),
    # so name the file by the last instant actually INCLUDED — otherwise
    # picking "to July 25" would produce a file that says July 26.
    inclusive_end = end - timedelta(microseconds=1)
    filename = (
        f"onboarding_{start.strftime('%Y-%m-%d')}"
        f"_to_{inclusive_end.strftime('%Y-%m-%d')}{scope}.xlsx"
    )
    return build_workbook(rows), filename
